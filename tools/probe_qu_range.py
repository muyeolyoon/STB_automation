"""Probe Q:U ART block on schedule spreadsheet (gid 205355322)."""
import os
import re
import sys
import zipfile

import openpyxl

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
sys.path.insert(0, os.path.join(ROOT, "stb-rpa"))

from component.schedule_loader import (  # noqa: E402
    download_drive_excel,
    worksheet_section_to_records,
    _find_section_bounds,
)

FILE_ID = "1fGc1yW9gBoHhJSo57E81FIAeAhNQz2ol"
GID = "205355322"
SA = os.path.join(ROOT, "stb-rpa", "service_account.json")
CACHE = os.path.join(ROOT, "test_log", "_schedule_cache", "probe_qu.xlsx")


def sheet_id_map(xlsx_path):
    with zipfile.ZipFile(xlsx_path) as z:
        xml = z.read("xl/workbook.xml").decode("utf-8", errors="replace")
    mapping = {}
    for m in re.finditer(r'<sheet[^>]+name="([^"]+)"[^>]+sheetId="(\d+)"', xml):
        mapping[m.group(2)] = m.group(1)
    for m in re.finditer(r'<sheet[^>]+sheetId="(\d+)"[^>]+name="([^"]+)"', xml):
        mapping[m.group(1)] = m.group(2)
    return mapping


def main():
    path = download_drive_excel(SA, FILE_ID, cache_path=CACHE)
    ids = sheet_id_map(path)
    print("sheetId -> name:")
    for sid, name in ids.items():
        mark = " <-- gid" if sid == GID else ""
        print(f"  {sid}: {name}{mark}")

    tab = ids.get(GID)
    if not tab:
        tab = "260709 모니터링"
        print(f"gid {GID} not in xlsx; fallback tab={tab}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[tab]
    print(f"\nTab: {tab}")
    for r in range(7, 15):
        row = list(
            ws.iter_rows(min_row=r, max_row=r, min_col=17, max_col=21, values_only=True)
        )[0]
        print(f"  row {r} Q-U:", row)

    rows = list(ws.iter_rows(values_only=True))
    hidx, cols = _find_section_bounds(rows, "uplus")
    print(f"\nART header row {hidx + 1}, cols:", {k: cols[k] for k in ("채널명", "채널번호", "광고편성 시간")})
    rec = worksheet_section_to_records(ws, "uplus")
    print(f"ART records: {len(rec)}")
    if rec:
        print("sample:", rec[:3])
    wb.close()


if __name__ == "__main__":
    main()
