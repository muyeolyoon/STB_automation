"""광고 편성표 로더 — Google Drive Excel 또는 gspread Google Sheet."""

import io
import os
import re
from datetime import datetime, time, timedelta

import openpyxl
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from oauth2client.service_account import ServiceAccountCredentials

from component.gspread_reader import load_sheet_data, sheet_tab_name

REQUIRED_COLUMNS = ("채널명", "채널번호", "광고편성 시간")
# 큐톤 모니터링 시트 — YYMMDD 모니터링 탭 ART(U+) / BigAD(SKB)
DEFAULT_DRIVE_FILE_ID = "1fGc1yW9gBoHhJSo57E81FIAeAhNQz2ol"
DEFAULT_GSPREAD_KEY = DEFAULT_DRIVE_FILE_ID
DRIVE_SHARED_OPTS = {"supportsAllDrives": True}
SECTION_MARKERS = {
    "skb": "BigAD",
    "bigad": "BigAD",
    "uplus": "ART",
    "art": "ART",
    "lgu": "ART",
}


def _drive_credentials(service_account_path):
    scope = [
        "https://www.googleapis.com/auth/drive.readonly",
        "https://www.googleapis.com/auth/spreadsheets.readonly",
    ]
    return ServiceAccountCredentials.from_json_keyfile_name(service_account_path, scope)


def default_monitor_tab_name():
    return f"{sheet_tab_name()} 모니터링"


def default_ad_schedule_tab_name():
    return f"{sheet_tab_name()} 광고편성"


def normalize_section(section=None):
    section = (section or os.environ.get("SCHEDULE_SECTION", "uplus")).strip().lower()
    if section in ("skb", "bigad"):
        return "skb"
    if section in ("uplus", "art", "lgu"):
        return "uplus"
    raise ValueError(
        f"알 수 없는 SCHEDULE_SECTION='{section}'. "
        "사용 가능: skb/bigad (BigAD), uplus/art/lgu (ART)"
    )


def section_label(section):
    section = normalize_section(section)
    return "BigAD(SKB)" if section == "skb" else "ART(U+)"


def _tab_candidates(sheet_name=None):
    yymmdd = sheet_tab_name()
    if sheet_name:
        return [sheet_name.strip()]
    return [
        default_monitor_tab_name(),
        yymmdd,
        f"{yymmdd}모니터링",
        datetime.now().strftime("%Y%m%d"),
        datetime.now().strftime("%Y-%m-%d"),
    ]


def resolve_worksheet_tab(workbook, sheet_name=None):
    names = workbook.sheetnames
    lowered = {name.strip().lower(): name for name in names}

    for candidate in _tab_candidates(sheet_name):
        key = candidate.strip().lower()
        if key in lowered:
            return lowered[key]

    yymmdd = sheet_tab_name()
    for name in names:
        compact = re.sub(r"\s+", "", name)
        if yymmdd in compact or yymmdd in name:
            return name

    raise ValueError(
        f"오늘 날짜({yymmdd})에 맞는 탭을 찾지 못했습니다. "
        f"탭 목록: {names[:20]}{'...' if len(names) > 20 else ''}. "
        f"SCHEDULE_SHEET_TAB 환경변수로 탭 이름을 지정하세요."
    )


def resolve_ad_schedule_tab(workbook):
    names = workbook.sheetnames
    lowered = {name.strip().lower(): name for name in names}
    yymmdd = sheet_tab_name()
    for candidate in (default_ad_schedule_tab_name(), f"{yymmdd}광고편성"):
        key = candidate.strip().lower()
        if key in lowered:
            return lowered[key]

    for name in names:
        compact = re.sub(r"\s+", "", name)
        if yymmdd in compact and "광고편성" in compact:
            return name
    return None


def _format_ad_time(value):
    if value is None or value == "":
        return ""

    if isinstance(value, datetime):
        return value.strftime("%H:%M:%S")

    if isinstance(value, time):
        return value.strftime("%H:%M:%S")

    if isinstance(value, timedelta):
        total = int(value.total_seconds())
        hours, rem = divmod(total, 3600)
        minutes, seconds = divmod(rem, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    if isinstance(value, (int, float)) and 0 <= float(value) < 1:
        total = int(round(float(value) * 86400))
        hours, rem = divmod(total, 3600)
        minutes, seconds = divmod(rem, 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    text = str(value).strip()
    if re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", text):
        parts = text.split(":")
        if len(parts) == 2:
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}:00"
        return f"{int(parts[0]):02d}:{int(parts[1]):02d}:{int(parts[2]):02d}"
    return text


def _normalize_channel_number(value):
    if value is None or value == "":
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str) and value.strip().isdigit():
        return int(value.strip())
    return value


def _normalize_row(row):
    normalized = {}
    for key, value in row.items():
        if key is None:
            continue
        header = str(key).strip()
        if not header:
            continue
        if isinstance(value, str):
            value = value.strip()
        normalized[header] = value

    if "광고편성 시간" in normalized:
        normalized["광고편성 시간"] = _format_ad_time(normalized["광고편성 시간"])
    if "채널번호" in normalized:
        normalized["채널번호"] = _normalize_channel_number(normalized["채널번호"])
    return normalized


MONITORING_DETAIL_TITLE = "■ 광고편성 모니터링"
# 모니터링 탭: BigAD=B열, ART=Q열 — 중간(O~P)은 여백
ART_BLOCK_MIN_COL = 12


def _section_end_col(row, title_col, marker):
    other = "ART" if marker == "BigAD" else "BigAD"
    for c_idx in range(title_col + 1, len(row)):
        cell = row[c_idx]
        if cell and other in str(cell):
            return c_idx
    return min(title_col + 14, len(row))


def _find_section_bounds(rows, section):
    section = normalize_section(section)
    marker = SECTION_MARKERS[section]
    title_col_idx = None
    anchor_row_idx = None

    # 1) Q7/B7 패턴 — "■ 광고편성 모니터링(상세)" 행으로 블록 열 고정 (ART=Q~, BigAD=B~)
    for r_idx, row in enumerate(rows[:35]):
        for c_idx, cell in enumerate(row):
            if not cell:
                continue
            text = str(cell).strip()
            if MONITORING_DETAIL_TITLE not in text:
                continue
            if section == "uplus" and c_idx >= ART_BLOCK_MIN_COL:
                title_col_idx = c_idx
                anchor_row_idx = r_idx
                break
            if section == "skb" and c_idx < ART_BLOCK_MIN_COL:
                title_col_idx = c_idx
                anchor_row_idx = r_idx
                break
        if title_col_idx is not None:
            break

    # 2) fallback: 상단 "ART 모니터링" / "BigAD 모니터링" 제목 행
    if title_col_idx is None:
        for r_idx, row in enumerate(rows[:25]):
            for c_idx, cell in enumerate(row):
                if cell and marker in str(cell):
                    anchor_row_idx = r_idx
                    title_col_idx = c_idx
                    break
            if title_col_idx is not None:
                break

    if title_col_idx is None:
        raise ValueError(
            f"시트에서 '{marker}' 모니터링 블록을 찾지 못했습니다 "
            f"(section={section_label(section)})."
        )

    start_row = anchor_row_idx if anchor_row_idx is not None else 0
    header_row_idx = None
    header_cols = {}
    for r_idx in range(start_row, min(start_row + 15, len(rows))):
        row = rows[r_idx]
        end_col = _section_end_col(row, title_col_idx, marker)
        segment = row[title_col_idx:end_col]
        local_headers = {}
        for offset, cell in enumerate(segment):
            if cell is None:
                continue
            header = str(cell).strip()
            if header:
                local_headers[header] = title_col_idx + offset
        if all(col in local_headers for col in REQUIRED_COLUMNS):
            header_row_idx = r_idx
            header_cols = local_headers
            break

    if header_row_idx is None:
        raise ValueError(
            f"'{marker}' 블록에서 필수 컬럼 {REQUIRED_COLUMNS} 헤더를 찾지 못했습니다."
        )

    return header_row_idx, header_cols


def worksheet_section_to_records(worksheet, section="uplus"):
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_row_idx, header_cols = _find_section_bounds(rows, section)
    records = []
    name_col = header_cols["채널명"]
    number_col = header_cols["채널번호"]
    time_col = header_cols["광고편성 시간"]

    for row in rows[header_row_idx + 1 :]:
        if not row:
            continue
        channel_name = row[name_col] if name_col < len(row) else None
        channel_number = row[number_col] if number_col < len(row) else None
        ad_time = row[time_col] if time_col < len(row) else None

        if (channel_name is None or str(channel_name).strip() == "") and (
            channel_number is None or str(channel_number).strip() == ""
        ):
            if records:
                break
            continue

        item = {}
        for header, col_idx in header_cols.items():
            item[header] = row[col_idx] if col_idx < len(row) else None
        item = _normalize_row(item)
        if not item.get("채널명") and not item.get("채널번호"):
            continue
        records.append(item)
    return records


def worksheet_ad_schedule_to_records(worksheet):
    """YYMMDD 광고편성 탭(wide matrix) → [{채널명, 채널번호, 광고편성 시간}, ...]."""
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = None
    slot_col_idx = None
    for idx, row in enumerate(rows[:25]):
        for c_idx, cell in enumerate(row):
            text = str(cell).strip() if cell is not None else ""
            if text != "구분":
                continue
            non_empty_after = sum(1 for v in row[c_idx + 1 :] if v not in (None, ""))
            if non_empty_after >= 5:
                header_idx = idx
                slot_col_idx = c_idx
                break
        if header_idx is not None:
            break
    if header_idx is None or header_idx + 2 >= len(rows):
        raise ValueError("'광고편성' 탭에서 채널명/채널번호 헤더 행을 찾지 못했습니다.")

    name_row = rows[header_idx + 1]
    number_row = rows[header_idx + 2]
    records = []
    for row in rows[header_idx + 3 :]:
        if not row:
            continue
        slot = row[slot_col_idx] if slot_col_idx < len(row) else None
        if slot in (None, ""):
            continue
        # 광고 회차 행만 사용. TOP/ALL 같은 구분 행은 제외.
        if isinstance(slot, str) and not re.fullmatch(r"\d+(\.0)?", slot.strip()):
            continue
        if isinstance(slot, (int, float)) and float(slot) <= 0:
            continue
        max_col = min(len(row), len(name_row), len(number_row))
        for col_idx in range(slot_col_idx + 1, max_col):
            channel_name = name_row[col_idx]
            channel_number = number_row[col_idx]
            ad_time = row[col_idx]
            if channel_name in (None, "") or channel_number in (None, ""):
                continue
            ad_time_text = _format_ad_time(ad_time)
            if not re.fullmatch(r"\d{1,2}:\d{2}(:\d{2})?", str(ad_time_text or "")):
                continue
            item = {
                "채널명": channel_name,
                "채널번호": channel_number,
                "광고편성 시간": ad_time_text,
            }
            records.append(_normalize_row(item))
    return records


def worksheet_to_records(worksheet, section=None):
    section = section or os.environ.get("SCHEDULE_SECTION")
    if section:
        return worksheet_section_to_records(worksheet, section=section)

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    if any(
        cell and ("BigAD" in str(cell) or "ART" in str(cell))
        for row in rows[:10]
        for cell in row
    ):
        return worksheet_section_to_records(worksheet, section=normalize_section())

    header_row_idx = None
    headers = None
    for idx, row in enumerate(rows[:30]):
        candidate = [str(cell).strip() if cell is not None else "" for cell in row]
        if all(col in candidate for col in REQUIRED_COLUMNS):
            header_row_idx = idx
            headers = candidate
            break

    if headers is None:
        raise ValueError(
            f"시트 '{worksheet.title}' 에 필수 컬럼 {REQUIRED_COLUMNS} 이 있는 헤더 행을 찾지 못했습니다."
        )

    records = []
    for row in rows[header_row_idx + 1 :]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            item[header] = row[idx] if idx < len(row) else None
        item = _normalize_row(item)
        if not item.get("채널명") and not item.get("채널번호"):
            continue
        records.append(item)
    return records


def list_drive_excel_tabs(service_account_path, drive_file_id):
    workbook = _load_workbook_from_drive(service_account_path, drive_file_id)
    names = workbook.sheetnames
    workbook.close()
    return names


def _parse_drive_modified_ts(modified_time: str) -> float:
    """Drive API modifiedTime(RFC3339) → epoch seconds."""
    if not modified_time:
        return 0.0
    text = modified_time.strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        return datetime.fromisoformat(text).timestamp()
    except ValueError:
        return 0.0


def download_drive_excel(service_account_path, drive_file_id, cache_path=None):
    creds = _drive_credentials(service_account_path)
    drive = build("drive", "v3", credentials=creds, cache_discovery=False)

    if cache_path and os.path.isfile(cache_path):
        try:
            meta = (
                drive.files()
                .get(
                    fileId=drive_file_id,
                    fields="modifiedTime",
                    **DRIVE_SHARED_OPTS,
                )
                .execute()
            )
            drive_mtime = _parse_drive_modified_ts(meta.get("modifiedTime", ""))
            cache_mtime = os.path.getmtime(cache_path)
            if drive_mtime and cache_mtime >= drive_mtime - 1:
                print(f"[schedule] 캐시된 편성표 사용: {cache_path}")
                return cache_path
            print("[schedule] Drive 편성표 갱신됨 — 캐시 재다운로드")
        except Exception as e:
            print(f"[schedule] 캐시 검증 실패({e}) — 재다운로드")

    meta = (
        drive.files()
        .get(fileId=drive_file_id, fields="name,mimeType", **DRIVE_SHARED_OPTS)
        .execute()
    )
    mime_type = meta.get("mimeType", "")

    if mime_type == "application/vnd.google-apps.spreadsheet":
        request = drive.files().export_media(
            fileId=drive_file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = drive.files().get_media(fileId=drive_file_id, **DRIVE_SHARED_OPTS)

    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()

    if cache_path:
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        with open(cache_path, "wb") as fh:
            fh.write(buffer.getvalue())
        print(f"[schedule] Drive Excel 저장: {cache_path} ({meta.get('name', drive_file_id)})")
        return cache_path

    tmp_path = os.path.join(
        os.environ.get("TEMP", os.environ.get("TMP", ".")),
        f"schedule_{drive_file_id[:8]}_{sheet_tab_name()}.xlsx",
    )
    with open(tmp_path, "wb") as fh:
        fh.write(buffer.getvalue())
    print(f"[schedule] Drive Excel 다운로드: {meta.get('name', drive_file_id)}")
    return tmp_path


def _load_workbook_from_drive(service_account_path, drive_file_id, cache_path=None):
    path = download_drive_excel(service_account_path, drive_file_id, cache_path=cache_path)
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def load_drive_excel_data(
    service_account_path,
    drive_file_id,
    sheet_name=None,
    cache_path=None,
    section=None,
):
    workbook = _load_workbook_from_drive(
        service_account_path,
        drive_file_id,
        cache_path=cache_path,
    )
    tab = resolve_worksheet_tab(workbook, sheet_name=sheet_name)
    section = normalize_section(section)
    print(f"[schedule] Drive Excel 탭: '{tab}' / {section_label(section)}")
    worksheet = workbook[tab]
    try:
        data = worksheet_to_records(worksheet, section=section)
    except ValueError as e:
        ad_tab = resolve_ad_schedule_tab(workbook)
        if "광고편성" not in str(tab) and ad_tab:
            print(
                f"[schedule] 모니터링 파싱 실패 → '{ad_tab}' wide 탭 fallback: {e}"
            )
            data = worksheet_ad_schedule_to_records(workbook[ad_tab])
        elif "광고편성" not in str(tab):
            workbook.close()
            raise
        else:
            print(
                f"[schedule] 모니터링 블록 파싱 실패 → 광고편성 wide 탭 fallback: {e}"
            )
            data = worksheet_ad_schedule_to_records(worksheet)
    if not data and "광고편성" not in str(tab):
        ad_tab = resolve_ad_schedule_tab(workbook)
        if ad_tab:
            print(f"[schedule] 모니터링 row 0건 → '{ad_tab}' wide 탭 fallback")
            data = worksheet_ad_schedule_to_records(workbook[ad_tab])
    workbook.close()
    print(f"[schedule] 편성 row 수: {len(data)}")
    return data


def load_local_excel_data(xlsx_path, sheet_name=None, section=None):
    if not os.path.isfile(xlsx_path):
        raise FileNotFoundError(f"편성표 파일 없음: {xlsx_path}")
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    tab = resolve_worksheet_tab(workbook, sheet_name=sheet_name)
    section = normalize_section(section)
    print(f"[schedule] 로컬 Excel 탭: '{tab}' / {section_label(section)}")
    worksheet = workbook[tab]
    try:
        data = worksheet_to_records(worksheet, section=section)
    except ValueError as e:
        ad_tab = resolve_ad_schedule_tab(workbook)
        if "광고편성" not in str(tab) and ad_tab:
            print(
                f"[schedule] 모니터링 파싱 실패 → '{ad_tab}' wide 탭 fallback: {e}"
            )
            data = worksheet_ad_schedule_to_records(workbook[ad_tab])
        elif "광고편성" not in str(tab):
            workbook.close()
            raise
        else:
            print(
                f"[schedule] 모니터링 블록 파싱 실패 → 광고편성 wide 탭 fallback: {e}"
            )
            data = worksheet_ad_schedule_to_records(worksheet)
    if not data and "광고편성" not in str(tab):
        ad_tab = resolve_ad_schedule_tab(workbook)
        if ad_tab:
            print(f"[schedule] 모니터링 row 0건 → '{ad_tab}' wide 탭 fallback")
            data = worksheet_ad_schedule_to_records(workbook[ad_tab])
    workbook.close()
    print(f"[schedule] 편성 row 수: {len(data)}")
    return data


def load_gspread_monitoring_data(
    service_account_path,
    spreadsheet_key=None,
    sheet_name=None,
    section=None,
):
    """Google Sheet YYMMDD 모니터링 탭 — SKB(A~C) / ART(E~G) 블록."""
    import gspread

    spreadsheet_key = spreadsheet_key or os.environ.get(
        "SPREADSHEET_KEY", DEFAULT_GSPREAD_KEY
    )
    sheet_name = sheet_name or default_monitor_tab_name()
    section = normalize_section(section)

    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
        "https://www.googleapis.com/auth/spreadsheets.readonly",
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_name(service_account_path, scope)
    client = gspread.authorize(creds)
    worksheet = client.open_by_key(spreadsheet_key).worksheet(sheet_name)
    rows = worksheet.get_all_values()

    data_col = 0 if section == "skb" else 4
    records = []
    for row in rows[2:]:
        if len(row) <= data_col + 2:
            continue
        channel_name = str(row[data_col] or "").strip()
        channel_number = row[data_col + 1] if data_col + 1 < len(row) else ""
        ad_time = str(row[data_col + 2] or "").strip()
        if not channel_name or not re.match(r"^\d{1,2}:\d{2}", ad_time):
            continue
        item = {
            "채널명": channel_name,
            "채널번호": _normalize_channel_number(channel_number),
            "광고편성 시간": _format_ad_time(ad_time),
        }
        records.append(_normalize_row(item))

    print(f"[schedule] gspread 모니터링 row 수: {len(records)} / {section_label(section)}")
    return records


def load_schedule_data(
    service_account_path,
    drive_file_id=None,
    spreadsheet_key=None,
    sheet_name=None,
    source=None,
    cache_dir=None,
    section=None,
):
    """편성표 로드. section: skb/bigad 또는 uplus/art/lgu (기본 uplus)."""
    source = (source or os.environ.get("SCHEDULE_SOURCE", "gspread")).strip().lower()
    drive_file_id = drive_file_id or os.environ.get(
        "DRIVE_SCHEDULE_FILE_ID", DEFAULT_DRIVE_FILE_ID
    )
    spreadsheet_key = spreadsheet_key or os.environ.get(
        "SPREADSHEET_KEY", DEFAULT_GSPREAD_KEY
    )
    sheet_name = sheet_name or os.environ.get("SCHEDULE_SHEET_TAB") or None
    local_xlsx = os.environ.get("SCHEDULE_XLSX_PATH", "").strip()
    section = normalize_section(section)

    if source == "local_xlsx":
        if not local_xlsx:
            raise ValueError("SCHEDULE_SOURCE=local_xlsx 이면 SCHEDULE_XLSX_PATH 가 필요합니다.")
        return load_local_excel_data(local_xlsx, sheet_name=sheet_name, section=section)

    if source == "gspread":
        if sheet_name is None:
            sheet_name = default_monitor_tab_name()
        print(f"[schedule] gspread 탭: '{sheet_name}' / {section_label(section)}")
        return load_gspread_monitoring_data(
            service_account_path,
            spreadsheet_key=spreadsheet_key,
            sheet_name=sheet_name,
            section=section,
        )

    cache_path = None
    if cache_dir:
        os.makedirs(cache_dir, exist_ok=True)
        cache_path = os.path.join(cache_dir, f"schedule_{sheet_tab_name()}.xlsx")

    return load_drive_excel_data(
        service_account_path,
        drive_file_id,
        sheet_name=sheet_name,
        cache_path=cache_path,
        section=section,
    )


def load_ad_schedule(
    service_account_path,
    section=None,
    drive_file_id=None,
    cache_dir=None,
    sheet_name=None,
):
    """cue 스크립트용: [{channel_name, channel, ad_time}, ...] 정렬 반환."""
    rows = load_schedule_data(
        service_account_path,
        drive_file_id=drive_file_id,
        sheet_name=sheet_name,
        cache_dir=cache_dir,
        section=section,
    )
    ad_schedule = []
    for row in rows:
        ad_time_str = row.get("광고편성 시간")
        channel = row.get("채널번호")
        channel_name = row.get("채널명")
        if not ad_time_str or channel in (None, ""):
            continue
        ad_time = datetime.strptime(ad_time_str, "%H:%M:%S")
        ad_dt = datetime.now().replace(
            hour=ad_time.hour,
            minute=ad_time.minute,
            second=ad_time.second,
            microsecond=0,
        )
        ad_schedule.append(
            {
                "channel_name": channel_name,
                "channel": channel,
                "ad_time": ad_dt,
            }
        )
    return sorted(ad_schedule, key=lambda x: x["ad_time"])
